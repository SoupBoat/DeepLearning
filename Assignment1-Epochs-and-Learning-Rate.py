#Assignment1
# import libraries
import os
import torch
import numpy as np
import torchvision.transforms as transforms
import torchvision.datasets as datasets
from torch.utils.data import DataLoader
import matplotlib.pyplot as plt
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import time
import openpyxl

print("importing complete")
#hyperparameters
# wb = openpyxl.load_workbook(r"C:\Users\deniz\OneDrive\Desktop\School\Yr4Sem2\ZEIT4154 Deep Learning\Assignments\Assignment 1\data.xlsx")    # for laptop
wb = openpyxl.load_workbook(r"/home/millerd/workspace/DeepLearning/data.xlsx")    # for remote desktop Will
# wb = openpyxl.load_workbook("data.xlsx")    # for remote desktop Andy
sheet = wb['Epochs']
for i in range(10, 30):
    n_epochs = sheet.cell(row = i+2, column = 1).value
    learning_rate = sheet.cell(row = i+2, column = 2).value
    lr_decay = sheet.cell(row = i+2, column = 3).value

    print("Epochs: ", n_epochs)
    print("Learning Rate: ", learning_rate)
    print("LR Decay: ", lr_decay)
    # n_epochs = 3 #make sure you continue training while the validation loss is reducing
    # learning_rate = 0.0001 #make sure your validation loss is reducing fairly smoothly each epoch
    batch_size = 128 #batch size is related to learning rate, if your batch size is larger you need a higher learning rate to update by the same amount each epoch, but the best values won't increase 1 to 1.
    lr_decay = 1 #this is the amount the learning rate will be reduced by each epoch, a lower learning rate is needed for fine tuning

    # # for google drive
    # traindir = r"/content/drive/MyDrive/Colab Notebooks/image_data/image_data/train"
    # testdir = r"/content/drive/MyDrive/Colab Notebooks/image_data/image_data/test"

    # for laptop
    # traindir = r"C:\Users\deniz\OneDrive\Desktop\School\Yr4Sem2\ZEIT4154 Deep Learning\Assignments\Assignment 1\image_data\image_data\train"
    # testdir = r"C:\Users\deniz\OneDrive\Desktop\School\Yr4Sem2\ZEIT4154 Deep Learning\Assignments\Assignment 1\image_data\image_data\test"                   #os.path.join(args.data, "test")

    # for remote desktop Andy
    # traindir = r"~/Documents/DL/test/image_data/train"
    # testdir = r"~/Documents/DL/test/image_data/test"

    # for remote desktop Will
    traindir = r"/home/millerd/workspace/DeepLearning/image_data/image_data/train"
    testdir = r"/home/millerd/workspace/DeepLearning/image_data/image_data/test"

    # normalize = transforms.Normalize(mean=[0.485, 0.456, 0.406],
    #                                  std=[0.229, 0.224, 0.225]) #these values are the rgb for imagenet.
    #                                 # It's okay to use these since the data set is from imagenet. could change it later if needed

    normalize = transforms.Normalize(mean=[0.485], std=[0.229]) #from imagenet as well. usually they use colour but this is for grayscale

    train_dataset = datasets.ImageFolder(
        traindir,
        transforms.Compose([
            # transforms.RandomResizedCrop(224), #This transformation performs a random cropping and resizing of the input image. 
            ## The 224 parameter specifies the target size of the cropped and resized image
            transforms.Grayscale(num_output_channels=1),
            transforms.Resize((224, 224)),  #images aren't the same size so resize.
            # transforms.RandomHorizontalFlip(),
            transforms.ToTensor(),
            normalize,  #if you normalize then the images become funky. if you want to visualise the images then comment this out
        ]))

    test_dataset = datasets.ImageFolder(
        testdir,
        transforms.Compose([
            # transforms.RandomResizedCrop(224), #This transformation performs a random cropping and resizing of the input image. 
            ## The 224 parameter specifies the target size of the cropped and resized image
            transforms.Grayscale(num_output_channels=1),
            transforms.Resize((224, 224)),  #images aren't the same size so resize.
            # transforms.RandomHorizontalFlip(),
            transforms.ToTensor(),
            normalize,
        ]))  

    # test if things work
    # print("train_dataset done")
    # print(train_dataset)
    # print(train_dataset.classes)
    # print(train_dataset.class_to_idx)

    train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)
    test_loader = DataLoader(test_dataset, batch_size=batch_size, shuffle=True)

    # data =  iter(train_loader)
    # images, labels = next(data)
    # fig = plt.figure(figsize=(15,5))
    # for idx in np.arange(20):
    #     ax = fig.add_subplot(4 , 5, idx+1,xticks=[],yticks=[])
    #     ax.imshow(images[idx].permute(1, 2, 0)) 
    #     ax.set_title(labels[idx].item())
    # plt.show()

    # define the deep model
    class Classifier(nn.Module):
        def __init__(self):
            super().__init__()
            self.fc1 = nn.Linear(1*224*224, 2048)    # 3*224*224 = 150k input units. 3 rgb then height width. 1*224*224 = 5000 for greyscale
            self.fc2 = nn.Linear(2048, 128)
            self.fc3 = nn.Linear(128,14)
            self.dropout = nn.Dropout(0.2)
        def forward(self,x):
            x =x.view(x.shape[0],-1).to(device)
            x =self.dropout(F.relu(self.fc1(x)))
            x =self.dropout(F.relu(self.fc2(x)))
            x = F.log_softmax(self.fc3(x), dim=1)
            return x

    device = 'cpu'   
    if torch.cuda.is_available():
        device = 'cuda'
    print('training on: ' + str(device))

    model = Classifier().to(device)
    loss_fn = nn.NLLLoss().to(device)        
    optimizer = optim.Adam(model.parameters(), lr=learning_rate)

    model.train()
    train_loss,test_losses = [],[]

    # for images, labels in train_loader:
    #     print("Train Labels:", labels.unique())

    # for images, labels in test_loader:
    #     print("Test Labels:", labels.unique())

    print("will funny")
    Start_time = time.time()
    #Training Step
    for epoch in range(n_epochs):
        running_loss = 0
        test_loss = 0
        correct = 0
        batch_count = 0
        for images,labels in train_loader:
            images, labels = images.to(device), labels.to(device)
            batch_count += 1
            if batch_count % 20 == 0:
                print(f"{round(100*batch_count/(17085/batch_size), 2)}% complete of training set") #17085 images
            optimizer.zero_grad()
            y_pred = model(images)
            loss = loss_fn(y_pred,labels).to(device)  
            loss.backward()
            optimizer.step()
            running_loss += loss.item()*images.size(0)
        for images,labels in test_loader:   #700 images
            images, labels = images.to(device), labels.to(device)
            batch_count += 1
            if batch_count % 20 == 0:
                print(f"{round(100*batch_count/(700/batch_size), 2)}% complete of training set") #700 images
            y_pred = model(images)
            loss = loss_fn(y_pred,labels).to(device)  
            test_loss += loss.item()*images.size(0)
            correct += (y_pred.argmax(1) == labels).type(torch.float).sum().item()

        learning_rate = learning_rate*lr_decay
        running_loss = running_loss/len(train_loader.sampler)
        test_loss = test_loss/len(test_loader.sampler)
        train_loss.append(running_loss)
        test_losses.append(test_loss)
        accuracy = 100* correct/len(test_dataset)

        sheet.cell(row = i+2, column = 4).value = accuracy
        sheet.cell(row = i+2, column = 5).value = running_loss
        sheet.cell(row = i+2, column = 6).value = test_loss
        sheet.cell(row = i+2, column = 7).value = round(time.time() - Start_time,2)
        if epoch % 10 == 0:
            wb.save("data.xlsx")
        print('Time: {:.2f} seconds'.format(time.time() - Start_time))
        print('Epoch: {} \tTraining Loss: {:.6f} \tValidation Loss: {:,.6f} \tValidation Accuracy: {:,.6f}'.format(epoch+1,running_loss,test_loss, accuracy))

    plt.subplot(3, 2, 1)
    plt.plot(train_loss, label='Train Loss SGD')
    plt.legend(loc='best')
    plt.subplot(3, 2, 2)
    plt.plot(test_losses, label = 'Test Loss SGD')
    plt.legend(loc='best')
    plt.show()
    print('Time: {:.2f} seconds'.format(time.time() - Start_time))

wb.save("data.xlsx")
wb.close()
