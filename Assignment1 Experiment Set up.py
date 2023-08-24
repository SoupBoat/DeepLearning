#Assignment1
# import libraries
import os
import torch
import numpy as np
import torchvision.transforms as transforms
import torchvision.datasets as datasets
from torch.utils.data import DataLoader
import matplotlib.pyplot as plt
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import time
import openpyxl

print("importing complete")

model = None    # global variable for the model. this is so that the model can be deleted and redefined for each experiment. hopefully this fixes some of the issues


# wb = openpyxl.load_workbook(r"C:\Users\deniz\OneDrive\Desktop\School\Yr4Sem2\ZEIT4154 Deep Learning\Assignments\Assignment 1\data.xlsx")    # for laptop
wbExperiments = openpyxl.load_workbook(r"/home/millerd/workspace/DeepLearning/Experiments.xlsx")    # for remote desktop Will
# wbExperiments = openpyxl.load_workbook(r"/home/deniz/Documents/DL/Experiments.xlsx")    # for ssh Andy

# wb = openpyxl.load_workbook("data.xlsx")    # for remote desktop Andy
ExpSheet = wbExperiments['Experiments']
# experiments = range(32, 38) #row corresponding to the experiments that you want to run. last number is not included. minimum is 2
experiments = [33]   # for the custom experiments

for experiment in experiments:
    for model_no in ["3"]: # models 2, 5, 6, 8, 9, 10 have been removed. do 7 after the greyscale ones are done
        wbModel = openpyxl.load_workbook("/home/millerd/workspace/DeepLearning/Model{}.xlsx".format(model_no))    # for remote desktop Will
        # wbModel = openpyxl.load_workbook("/home/deniz/Documents/DL/Model{}.xlsx".format(model_no))      # for ssh Andy
        modelSheet = wbModel['Results']
        input_size = modelSheet.cell(row = 3, column = 1).value
        hidden_size1 = modelSheet.cell(row = 3, column = 2).value
        hidden_size2 = modelSheet.cell(row = 3, column = 3).value
        output_size = modelSheet.cell(row = 3, column = 4).value

        # get the values from the excel sheet/reset the values
        experiment_type = ExpSheet.cell(row = experiment, column = 1).value
        n_epochs = ExpSheet.cell(row = experiment, column = 2).value
        batch_size = ExpSheet.cell(row = experiment, column = 3).value
        learning_rate = ExpSheet.cell(row = experiment, column = 4).value
        # defaults
        momentum = 1
        lr_decay = 1
        weight_decay = 0
        dropout_prob = 0.2
        
        if experiment_type == "Learning Rate Decay":
            sheet = wbModel["Learning Rate Decay"]
            lr_decay = ExpSheet.cell(row = experiment, column = 6).value   # in the excel sheet, column 5 is the learning rate decay
        elif experiment_type == "Momentum":
            sheet = wbModel["Momentum"]
            momentum = ExpSheet.cell(row = experiment, column = 5).value   # in the excel sheet, column 5 is the momentum
        elif experiment_type == "Batch Size":
            sheet = wbModel["Batch Size"]
        elif experiment_type == "Learning Rate":
            sheet = wbModel["Learning Rate"]
        elif experiment_type == "Weight Decay":
            sheet = wbModel["Weight Decay"]
            weight_decay = ExpSheet.cell(row = experiment, column = 7).value   # in the excel sheet, column 5 is the weight decay
        elif experiment_type == "Dropout":
            sheet = wbModel["Dropout"]
            dropout_prob = ExpSheet.cell(row = experiment, column = 8).value
        elif experiment_type == "Custom":
            sheet = wbModel["Custom"]
            momentum = ExpSheet.cell(row = experiment, column = 5).value
            lr_decay = ExpSheet.cell(row = experiment, column = 6).value
            weight_decay = ExpSheet.cell(row = experiment, column = 7).value
            dropout_prob = ExpSheet.cell(row = experiment, column = 8).value
        
        print("Experiment Type: ", experiment_type)
        print("Epochs: ", n_epochs)
        print("Learning Rate: ", learning_rate)
        print("LR Decay: ", lr_decay)
        print("Batch Size: ", batch_size)
        print("Momentum: ", momentum)
        print("Weight Decay: ", weight_decay)
        print("Dropout: ", dropout_prob)
        print("Experiment Row: ", experiment)
        print("Model: ", model_no)

        # n_epochs = 3 #make sure you continue training while the validation loss is reducing
        # learning_rate = 0.0001 #make sure your validation loss is reducing fairly smoothly each epoch
        # batch_size = 128 #batch size is related to learning rate, if your batch size is larger you need a higher learning rate to update by the same amount each epoch, but the best values won't increase 1 to 1.
        # lr_decay = 0.5 #this is the amount the learning rate will be reduced by each epoch, a lower learning rate is needed for fine tuning

        # # for google drive
        # traindir = r"/content/drive/MyDrive/Colab Notebooks/image_data/image_data/train"
        # testdir = r"/content/drive/MyDrive/Colab Notebooks/image_data/image_data/test"

        # for laptop
        # traindir = r"C:\Users\deniz\OneDrive\Desktop\School\Yr4Sem2\ZEIT4154 Deep Learning\Assignments\Assignment 1\image_data\image_data\train"
        # testdir = r"C:\Users\deniz\OneDrive\Desktop\School\Yr4Sem2\ZEIT4154 Deep Learning\Assignments\Assignment 1\image_data\image_data\test"                   #os.path.join(args.data, "test")

        # # for remote desktop Andy
        # traindir = r"~/Documents/DL/test/image_data/train"
        # testdir = r"~/Documents/DL/test/image_data/test"

        # for ssh Andy
        # traindir = r"~/Documents/DL/image_data/train"
        # testdir = r"~/Documents/DL/image_data/test"

        # for remote desktop Will
        traindir = r"/home/millerd/workspace/DeepLearning/image_data/image_data/train"
        testdir = r"/home/millerd/workspace/DeepLearning/image_data/image_data/test"

        # if input_size == 3*224*224:
        #     normalize = transforms.Normalize(mean=[0.485, 0.456, 0.406],
        #                                  std=[0.229, 0.224, 0.225]) #these values are the rgb for imagenet.
        #         # It's okay to use these since the data set is from imagenet. could change it later if needed
        # else:
        normalize = transforms.Normalize(mean=[0.485], std=[0.229]) #from imagenet as well. usually they use colour but this is for grayscale


        # preprocessTrain = []
        # if input_size == 1*224*224:
        #     preprocessTrain.append(transforms.Grayscale(num_output_channels=1))  #converts to grayscale
        # preprocessTrain.extend([
        #     transforms.RandomErasing(), #randomly erases parts of the image
        #     transforms.RandomResizedCrop(224), #randomly crop the image to 224x224
        #     # transforms.ColorJitter(brightness=0.1, contrast=0.1, saturation=0.1, hue=0.1), #randomly change the brightness, contrast, saturation and hue
        #     # transforms.Resize((224, 224)),  #images aren't the same size so resize.
        #     transforms.RandomHorizontalFlip(),  # randomly flip and rotate. 50% chance of flipping
        #     transforms.ToTensor(),
        #     normalize   #normalizes the tensor with the mean and std of the imagenet data set
        # ])

        preprocessTrain = []
        # if input_size == 1*224*224:
        #     preprocessTrain.append(transforms.Grayscale(num_output_channels=1))  #converts to grayscale
        preprocessTrain.extend([
            # transforms.RandomErasing(), #randomly erases parts of the image
            # transforms.RandomResizedCrop(224), #randomly crop the image to 224x224
            # transforms.ColorJitter(brightness=0.5, contrast=0.5, saturation=0.5, hue=0.5), #randomly change the brightness, contrast, saturation and hue
            transforms.Grayscale(num_output_channels=1),
            transforms.Resize((224, 224)),  #images aren't the same size so resize.
            transforms.RandomHorizontalFlip(),  # randomly flip and rotate. 50% chance of flipping
            transforms.ToTensor(),
            normalize   #normalizes the tensor with the mean and std of the imagenet data set
            # transforms.RandomErasing() #randomly erases parts of the image
        ])

        for process in preprocessTrain:
            print(process)

        train_dataset = datasets.ImageFolder(
            traindir,
            transforms.Compose(preprocessTrain))

        test_dataset = datasets.ImageFolder(
            testdir,
            transforms.Compose([
            transforms.Grayscale(num_output_channels=1),
            transforms.Resize((224, 224)),  #images aren't the same size so resize.
            transforms.ToTensor(),
            normalize   #normalizes the tensor with the mean and std of the imagenet data set
        ]))  



        train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)
        test_loader = DataLoader(test_dataset, batch_size=batch_size, shuffle=True)

        classes = train_dataset.classes
        print()

        # class for defining the model
        class Classifier(nn.Module):
            def __init__(self, input_size, hidden_size1, hidden_size2, output_size, dropout_prob):
                super(Classifier, self).__init__()
                self.fc1 = nn.Linear(input_size, hidden_size1)
                self.fc2 = nn.Linear(hidden_size1, hidden_size2)
                self.fc3 = nn.Linear(hidden_size2, output_size)
                self.dropout = nn.Dropout(dropout_prob)
            def forward(self, x):
                x = x.view(x.shape[0], -1).to(device)
                x = self.dropout(F.relu(self.fc1(x)))
                x = self.dropout(F.relu(self.fc2(x)))
                # x = F.log_softmax(self.fc3(x), dim=1)
                x = self.fc3(x)
                return x

        device = 'cpu'   
        if torch.cuda.is_available():
            device = 'cuda'
        print('training on: ' + str(device))

        # Close and delete the existing model and optimizer. hopefully this fixes some of the issues
        if model is not None:
            del model
            del optimizer
            torch.cuda.empty_cache()  # Release GPU memory

        # create model from the values given in the excel sheet
        model = Classifier(input_size=input_size, hidden_size1=hidden_size1, hidden_size2=hidden_size2, output_size=14, dropout_prob=dropout_prob).to(device)

        loss_fn = nn.CrossEntropyLoss().to(device)  
        # loss_fn = nn.CrossEntropyLoss().to(device)      
        optimizer = optim.SGD(model.parameters(), lr=learning_rate, momentum=momentum, weight_decay=weight_decay)

        model.train()
        train_loss,test_losses = [],[]

        # for images, labels in train_loader:
        #     print("Train Labels:", labels.unique())

        # for images, labels in test_loader:
        #     print("Test Labels:", labels.unique())



        Start_time = time.time()
        try:   #if the batch size is larger than the hidden layer size, the model will not train
            #Training Step
            for epoch in range(n_epochs):
                running_loss = 0
                test_loss = 0
                correct = 0
                batch_count = 0
                for images, labels in train_loader:
                    images, labels = images.to(device), labels.to(device)
                    batch_count += 1
                    if batch_count % 20 == 0:
                        print(f"{round(100*batch_count/(17085/batch_size), 2)}% complete of training set") #17085 images
                    optimizer.zero_grad()
                    y_pred = model(images)
                    loss = loss_fn(y_pred,labels).to(device)  
                    loss.backward()
                    optimizer.step()
                    running_loss += loss.item()*images.size(0)
                batch_count = 0
                with torch.no_grad():
                    for images,labels in test_loader:   #700 images
                        images, labels = images.to(device), labels.to(device)
                        batch_count += 1
                        if batch_count % 20 == 0:
                            print(f"{round(100*batch_count/(700/batch_size), 2)}% complete of training set") #700 images
                        y_pred = model(images)
                        loss = loss_fn(y_pred,labels).to(device)  
                        test_loss += loss.item()*images.size(0)
                        correct += (y_pred.argmax(1) == labels).type(torch.float).sum().item()

                    # prepare to count predictions for each class
                    correct_pred = {classname: 0 for classname in classes}
                    total_pred = {classname: 0 for classname in classes}   

                    # again no gradients needed
                    with torch.no_grad():
                        for data in test_loader:
                            images, labels = data    
                            images, labels    = images.to(device), labels.to(device)
                            outputs = model(images)    
                            _, predictions = torch.max(outputs, 1)
                            # collect the correct predictions for each class
                            for label, prediction in zip(labels, predictions):
                                if label == prediction:
                                    correct_pred[classes[label]] += 1
                                total_pred[classes[label]] += 1

                    total_correct = 0
                    classAccuracy = []
                    # print accuracy for each class
                    for classname, correct_count in correct_pred.items():
                        accuracy = 100 * float(correct_count) / total_pred[classname]
                        total_correct += correct_count
                        print("Accuracy for class {:5s} is: {:.2f} %".format(classname, 
                                                                    accuracy))
                        classAccuracy.append(round(accuracy,2))

                    print("Final accuracy is: {:.2f}%".format((total_correct/len(test_dataset))*100))

                
                    running_loss = running_loss/len(train_loader.sampler)
                    test_loss = test_loss/len(test_loader.sampler)
                    train_loss.append(running_loss)
                    test_losses.append(test_loss)
                    accuracy = 100 * correct/len(test_dataset)
                    # Save values into spreadsheet for tracking
                    next_row = sheet.max_row + 1
                    sheet.cell(row = next_row, column = 1).value = epoch + 1
                    sheet.cell(row = next_row, column = 2).value = learning_rate
                    sheet.cell(row = next_row, column = 3).value = lr_decay
                    sheet.cell(row = next_row, column = 4).value = batch_size
                    sheet.cell(row = next_row, column = 5).value = momentum
                    sheet.cell(row = next_row, column = 6).value = weight_decay
                    sheet.cell(row = next_row, column = 7).value = accuracy
                    sheet.cell(row = next_row, column = 8).value = running_loss
                    sheet.cell(row = next_row, column = 9).value = test_loss
                    sheet.cell(row = next_row, column = 10).value = round(time.time() - Start_time,2)
                    #add all of the class accuracies to the spreadsheet in the next column
                    for i in range(len(classAccuracy)):
                        sheet.cell(row = next_row, column = 11+i).value = classAccuracy[i]

                    learning_rate = learning_rate*lr_decay
                    wbModel.save("Model{}.xlsx".format(model_no))
                    print('Time: {:.2f} seconds'.format(time.time() - Start_time))
                    print('Epoch: {} \tTraining Loss: {:.6f} \tValidation Loss: {:,.6f} \tValidation Accuracy: {:,.6f}'.format(epoch+1,running_loss,test_loss, accuracy))
                    
                    
                    # Plot the training and validation losses on the same graph for easy comparison
                    plt.figure()
                    plt.plot(train_loss, label='Train Loss SGD')
                    plt.plot(test_losses, label = 'Test Loss SGD')
                    plt.legend(loc='best')
                    plt.title("Model{} ExpType{} ExpRow{}".format(model_no, experiment_type, experiment))
                    # plt.show()
                    # save the plots in the folder corresponding the model number
                    plt.savefig("Figures/Model{}/Model{}_ExpType{}_ExpRow{}.png".format(model_no, model_no, experiment_type, experiment))
        except:
            next_row = sheet.max_row + 1
            sheet.cell(row = next_row, column = 1).value = "Error"
            print("Error")

        wbModel.save("Model{}.xlsx".format(model_no))
        wbModel.close()
        torch.cuda.empty_cache()